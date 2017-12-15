#!/usr/bin/python

import re

import os

import openpyxl

import glob

import csv

import xlrd


from openpyxl import Workbook

path ='/home/chris/scrpits/CustomerForms/*.txt'

filelist = glob.glob(path)

contents = ' '
for file in filelist:
	try:
		open_file= open(file,'r')
	except:
		print"%s failed to open" % file
	contents = contents + open_file.read() + "\n"

	try:
		open_file.close()
	except:
		print"file closed"

phoneRegex = re.compile(r'''(
	(\d{3}|\(\d{3}\))? # area code
	(\s|-|\.)? # seperator
	(\d{3}) # first 3 digits
	(\s|-|\.) # seperator
	(\d{4}) # last 4 digits
	(\s*(ext|x|ext.)\s*(\d{2,5}))? # extension
	)''', re.VERBOSE)

emailRegex = re.compile(r'''(
	[a-zA-Z0-9._%+-]+
	@
	[a-zA-Z0-9.-]+
	(\.[a-zA-Z]{2,4})
	)''', re.VERBOSE)

firstNameRegex = re.compile(r'''(
	(First\sName:)
	(\s)?
	(([A-Z]{1})
	(\')?
	([A-Z}{1})?
	([a-z]+))
	(\s)
	(\n)
	)''', re.VERBOSE)

lastNameRegex = re.compile(r'''(
	(Last\sName:)
	(\s)+
	(([A-Z}{1})
	([a-z]+)
	(\-)?
	([A-Z}{1})?
	([a-z]+)?)
	(\s)
	(\n)
	)''', re.VERBOSE)

addressRegex = re.compile(r'''(
	([A-Z]{2})
	(\s)+
	(\d{5})
	(\,|-)*
	(\d{4})?
	)''', re.VERBOSE)

courseRegex = re.compile(r'''(
	(Course\sName:)
	(\s)?
	([A-Za-z\s]+)
	(\s)
	(\n)
	)''', re.VERBOSE)

courses = []
phone = []
email = []
firstName = []
lastName = []
address = []
total = []



for groups in phoneRegex.findall(contents):
	phoneNum = '-'.join([groups[1], groups[3], groups[5]])
	phone.append(phoneNum)
	total.append(phoneNum)

for groups in emailRegex.findall(contents):
	email.append(groups[0])
	total.append(groups[0])

for groups in firstNameRegex.findall(contents):
	firstName.append(groups[3])
	total.append(groups[3])

for groups in lastNameRegex.findall(contents):
	lastName.append(groups[3])
	total.append(groups[3])

for groups in addressRegex.findall(contents):
	address.append(groups[0])
	total.append(groups[0])

for groups in courseRegex.findall(contents):
	courses.append(groups[3])
	total.append(groups[3])


if len(total) > 0:
	print('\n'.join(total))

	jim = openpyxl.Workbook()
	AS = jim.create_sheet(index = 0, title = "Results")
	AS['A1'] = "Course Name"
	AS['D1'] = "Email"
	AS['F1'] = "ZipCodes"
	AS['B1'] = "First Name"
	AS['C1'] = "Last Name"
	AS['E1'] = "Phone Number"


	index = 2
	for match in phone:
		cellobj = AS.cell(row = index, column = 5)
		cellobj.value = match
		index = index + 1

	index = 2
	for match in  email:
		cellobj = AS.cell(row = index, column = 4)
		cellobj.value = match
		index = index + 1

	index = 2
	for match in  address:
		cellobj = AS.cell(row = index, column = 6)
		cellobj.value = match
		index = index + 1

	index = 2
	for match in  firstName:
		cellobj = AS.cell(row = index, column = 2)
		cellobj.value = match
		index = index + 1

	index = 2
	for match in  lastName:
		cellobj = AS.cell(row = index, column = 3)
		cellobj.value = match
		index = index + 1

	index = 2
	for match in  courses:
		cellobj = AS.cell(row = index, column = 1)
		cellobj.value = match
		index = index + 1

	jim.save('example2_test.xlsx')



import xlrd
import csv

with xlrd.open_workbook('example2_test.xlsx') as wb:
    sh = wb.sheet_by_index(0)  # or wb.sheet_by_name('name_of_the_sheet_here')
    with open('test.csv', 'wb') as f:
        c = csv.writer(f)
        for r in range(sh.nrows):
            c.writerow(sh.row_values(r))

